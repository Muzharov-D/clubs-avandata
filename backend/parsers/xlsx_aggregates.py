"""Командные агрегаты СУММОЙ из per-player CSV/XLSX (двуязычные заголовки).

Зачем: часть экспортов SportVisor приходит урезанным PDF (25 стр.) без детальных
команд-страниц → team_aggregates из PDF пустые. Но per-player xlsx/csv содержит
все нужные счётные колонки — суммируем их в командные агрегаты.

Выход — ТА ЖЕ структура, что derivedAggregates в upload/routes.ts:
  { shooting:{shots:{value}, ...}, passes:{...}, ... }  (только {value}).
Мерджится там по приоритету: build_match(PDF) > xlsx-derived > rich-derived.

Заголовки колонок бывают РУССКИЕ («удары», «пас») и АНГЛИЙСКИЕ («shot», «pass»)
— карта знает оба. Числа в xlsx чистые (умножение цифр — артефакт ТЕКСТА PDF,
не xlsx).

Usage: python xlsx_aggregates.py <input.xlsx|csv> <output.json>
"""
import argparse, json, sys, csv as csvmod, io, re

# Метрика агрегата → (русское имя колонки, английское имя колонки).
# Сопоставление по точному совпадению заголовка (lower, trim).
SHOOTING = {
    'shots':    ('удары', 'shot'),
    'onTarget': ('точные удары', 'shot_success'),
    'goals':    ('голы', 'goal'),
    'byHead':   ('удары головой', 'headshot'),
}
PASSES = {
    'total':        ('пас', 'pass'),
    'successful':   ('удачные пасы', 'pass_success'),
    'progressive':  ('прогрессивные пасы', 'progressive_pass'),
    'toFinalThird': ('пасы в финальную треть', 'pass_to_final_third'),
    'intoPenArea':  ('пасы в штрафную', 'pass_in_pen_area'),
    'crosses':      ('кроссы', 'cross'),
    'keyPass':      ('ключевые пасы', 'key_pass'),
}
ATTACKS = {
    'assists':      ('голевые передачи', 'assist'),
    'goalActions':  ('передачи под удар в створ', 'shot_assist'),
    'dribble':      ('дриблинг', 'dribble'),
    'touchesInBox': ('касания мяча в штрафной', 'ball_touch_pen_area'),
    'entriesInBox': ('входы в штрафную', 'entries_in_penalty_area'),
}
POSSESSION = {
    'lostBall':         ('потери мяча', 'lost_ball'),
    'technicalMistake': ('технические ошибки', 'technical_mistake'),
    'loseOnOwnHalf':    ('потери на своей половине', 'lost_ball_own_half'),
}
RECOVERIES = {
    'tackle':       ('отборы', 'tackle'),
    'interception': ('перехваты', 'interception'),
    'recovery':     ('возвраты', 'recovery'),
}
DUELS = {
    'duel':       ('дуэли', 'duel'),
    'aerialDuel': ('воздушные дуэли', 'ariel_duel'),  # 'ariel' — опечатка в экспорте
}
PRESSING = {
    'pressing':        ('прессинг', 'pressing'),
    'counterpressing': ('контрпрессинг', 'contrpressing'),  # 'contr' — как в экспорте
}
POSITIONING = {
    'clearance':   ('выносы', 'clearance'),
    'rebounds':    ('подборы', 'rebounds'),
    'blockedShot': ('заблокированные удары', 'blocked_shot'),
}
SET_PIECES = {
    'corner':       ('угловые', 'corner'),
    'freeKickShot': ('штрафные с ударом', 'free_kick_with_shot'),
    'freeKick':     ('штрафные удары', 'free_kick'),
    'penalty':      ('пенальти', 'penalty'),
    'throwing':     ('ауты', 'throwing'),
}
SECTIONS = {
    'shooting': SHOOTING, 'passes': PASSES, 'attacks': ATTACKS,
    'possession': POSSESSION, 'recoveriesAndTackling': RECOVERIES,
    'duels': DUELS, 'pressing': PRESSING, 'positioning': POSITIONING,
    'setPieces': SET_PIECES,
}
# Фитнес-метрики команды (сумма дистанций) — отдельная секция, для полноты.
FITNESS = {
    'totalDistance':  ('общая дистанция', 'distance_total'),
    'sprintDistance': ('дистанция спринтов', 'sprints_distance'),
}


def to_num(s):
    if s is None:
        return 0.0
    t = str(s).strip().replace(',', '.')
    if t in ('', '-'):
        return 0.0
    try:
        return float(t)
    except ValueError:
        return 0.0


def read_rows(path):
    """→ (headers_lower, list[dict row by header_lower]). Поддержка xlsx и csv."""
    if path.lower().endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb['Player Report'] if 'Player Report' in wb.sheetnames else wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        with open(path, encoding='utf-8-sig') as f:
            lines = [ln for ln in f.readlines() if not ln.lstrip().startswith('#')]
        grid = [r for r in csvmod.reader(io.StringIO(''.join(lines)))]
    # Найти строку-заголовок: содержит 'номер'/'jersey no.'
    hi = None
    for i, r in enumerate(grid):
        cells = [str(c).strip().lower() if c is not None else '' for c in r]
        if 'номер' in cells or 'jersey no.' in cells:
            hi = i
            break
    if hi is None:
        return [], []
    headers = [str(c).strip().lower() if c is not None else '' for c in grid[hi]]
    rows = []
    num_keys = ('номер', 'jersey no.')
    ni = next((i for i, h in enumerate(headers) if h in num_keys), None)
    for r in grid[hi + 1:]:
        if ni is None or ni >= len(r) or r[ni] in (None, ''):
            continue
        rows.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return headers, rows


def col_index(headers, ru, en):
    for name in (en, ru):
        if name and name in headers:
            return name
    return None


def sum_col(rows, key):
    return sum(to_num(row.get(key)) for row in rows) if key else 0.0


def fmt(v):
    return int(v) if float(v).is_integer() else round(v, 2)


def build(path):
    headers, rows = read_rows(path)
    if not rows:
        return {'players': 0, 'aggregates': {}}
    out = {}
    for section, metrics in SECTIONS.items():
        out[section] = {}
        for metric, (ru, en) in metrics.items():
            key = col_index(headers, ru, en)
            out[section][metric] = {'value': fmt(sum_col(rows, key))}
    # fitness — суммы дистанций (метры)
    out['fitness'] = {}
    for metric, (ru, en) in FITNESS.items():
        key = col_index(headers, ru, en)
        out['fitness'][metric] = {'value': fmt(sum_col(rows, key))}
    return {'players': len(rows), 'aggregates': out}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('input')
    ap.add_argument('output')
    args = ap.parse_args()
    data = build(args.input)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"OK players={data['players']} sections={len(data['aggregates'])}", file=sys.stderr)


if __name__ == '__main__':
    main()
