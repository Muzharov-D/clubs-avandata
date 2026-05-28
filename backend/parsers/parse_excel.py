"""Parse SportVisor Excel export (Player Report sheet → 136+ columns per-player).

Usage:
  python parse_excel.py <input.xlsx> <output.json>

Output JSON:
  {
    "match": { homeTeam, awayTeam, date, score? },
    "players": [
      { number, name, minutes,
        stats: { passing:{...}, attacking:{...}, defending:{...}, fitness:{...}, ... },
        raw:   { '<header>': value, ... }
      }, ...
    ]
  }

Группировка 136 колонок — по русским словам в headers:
  pas/peredach → passing; udar/gol/xg → attacking; otbor/perehv → defending;
  dribling → dribbling; distanc/intensivn → fitness; press → pressing;
  угло/штрафн/аут → setpieces; нарушен → fouls.
"""
import argparse, json, re
from openpyxl import load_workbook


def slugify(s):
    if not s: return "unknown"
    s = str(s).strip().lower()
    table = str.maketrans({
        'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e','ж':'zh',
        'з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n','о':'o',
        'п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h','ц':'c',
        'ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya',
    })
    s = s.translate(table)
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s or "unknown"


GROUPS = [
    (['пас', 'передач'],              'passing'),
    (['удар', 'гол', 'xg', 'xa'],     'attacking'),
    (['перехват', 'защит'],           'defending'),
    (['дриблинг', 'обводк'],          'dribbling'),
    (['дистанц', 'интенсив', 'спринт', 'км', 'бег'], 'fitness'),
    (['отбор', 'единоборств'],        'duels'),
    (['нарушен', 'фол'],              'fouls'),
    (['угло', 'штрафн', 'аут'],       'setpieces'),
    (['прессинг'],                    'pressing'),
]


def categorize(header):
    lower = (header or '').lower()
    for prefixes, group in GROUPS:
        for p in prefixes:
            if p in lower:
                return group
    return 'other'


def parse(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = 'Player Report' if 'Player Report' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    meta = {}
    for r in range(1, 5):
        cell = ws.cell(row=r, column=1).value
        if not cell: continue
        m = re.match(r'^([^:]+):\s*(.+)$', str(cell).strip())
        if m:
            meta[m.group(1).strip().lower()] = m.group(2).strip()

    headers = [str(ws.cell(row=6, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]

    players = []
    for r in range(7, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None for v in row): continue
        number = row[0]
        name   = row[1]
        if number is None and name is None: continue

        raw = {}
        grouped = {}
        for h, v in zip(headers[2:], row[2:]):
            if not h: continue
            raw[h] = v
            cat = categorize(h)
            slug = slugify(h)
            grouped.setdefault(cat, {})[slug] = v

        minutes = None
        for h, v in zip(headers, row):
            if h and 'игровое время' in h.lower():
                try: minutes = int(v) if v is not None else None
                except (ValueError, TypeError): minutes = None
                break

        players.append({
            'number': str(number).zfill(2) if number is not None else '',
            'name':   str(name) if name else '',
            'minutes': minutes,
            'stats': grouped,
            'raw': raw,
        })

    score = None
    for key in ('result', 'счет', 'счёт'):
        if key in meta:
            m = re.match(r'(\d+)\s*[:-]\s*(\d+)', meta[key])
            if m: score = {'home': int(m.group(1)), 'away': int(m.group(2))}; break

    return {
        'match': {
            'homeTeam':     meta.get('home team') or meta.get('team'),
            'awayTeam':     meta.get('guest team'),
            'date':         meta.get('date'),
            'score':        score,
            'sourceSheet':  sheet_name,
            'columnsCount': len(headers),
        },
        'players': players,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('input_xlsx')
    ap.add_argument('output_json')
    args = ap.parse_args()

    data = parse(args.input_xlsx)
    with open(args.output_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"OK columns={data['match']['columnsCount']} players={len(data['players'])}")


if __name__ == '__main__':
    main()
