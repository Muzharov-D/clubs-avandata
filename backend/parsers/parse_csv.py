"""Parse SportVisor/Наградион per-player CSV export → JSON в том же неймспейсе,
что и parse_zenit_full.py (группы attack/defence/fitness, английские ключи).

Формат входа (UTF-8):
  #Team: ...
  #Home team: ...
  #Guest Team: ...
  #Date: ...
  номер,имя,игровое время,<~130 русских колонок>
  5,Михаил Галицкий,82.0,46,37,80.4,...
  ...

Зачем: PDF-парсер извлекает те же числа, но его табличная экстракция хрупкая
(порядок колонок). CSV — структурированный эталон, поэтому служит источником
истины в мердже (см. upload/routes.ts). Компаунды собираются из троек
«значение, удачные, точность» в {total, successful, accuracy}.

Usage:
  python parse_csv.py <input.csv> <output.json>
"""
import argparse, csv, io, json, re, sys

# (group, key, total_col, successful_col|None, accuracy_col|None)
# Имена колонок — точно как в шапке CSV. Компаунд собирается если задан succ/acc.
MAPPING = [
    # ── attack: passing ──
    ('attack', 'pass',             'пас', 'удачные пасы', 'точность пасов'),
    ('attack', 'passForward',      'пасы вперед', 'удачные пасы вперед', 'точность пасов вперед'),
    ('attack', 'passBack',         'пасы назад', 'удачные пасы назад', 'точность пасов назад'),
    ('attack', 'passSideways',     'пасы поперек', 'удачные пасы поперек', 'точность пасов поперек'),
    ('attack', 'passShort',        'короткие пасы', 'удачные короткие пасы', 'точность коротких пасов'),
    ('attack', 'passMiddle',       'средние пасы', 'удачные средние пасы', 'точность средних пасов'),
    ('attack', 'passLong',         'длинные пасы', 'удачные длинные пасы', 'точность длинных пасов'),
    ('attack', 'intoPenArea',      'пасы в штрафную', 'удачные пасы в штрафную', 'точность пасов в штрафную'),
    ('attack', 'passToFinalThird', 'пасы в финальную треть', 'удачные пасы в финальную треть', 'точность пасов в финальную треть'),
    ('attack', 'progressivePass',  'прогрессивные пасы', 'удачные прогрессивные пасы', 'точность прогрессивных пасов'),
    ('attack', 'throughPass',      'разрезающие пасы', 'удачные разрезающие пасы', 'точность разрезающих пасов'),
    ('attack', 'keyPass',          'ключевые пасы', None, None),
    ('attack', 'cross',            'кроссы', 'удачные кроссы', 'точность кроссов'),
    ('attack', 'assist',           'голевые передачи', None, None),
    ('attack', 'secondAssist',     'предголевые передачи', None, None),
    ('attack', 'thirdAssist',      'третьи передачи (third assist)', None, None),
    ('attack', 'passOnTarget',     'передачи под удар в створ', None, None),
    ('attack', 'receivedPass',     'принятые пасы', None, None),
    ('attack', 'entriesInBox',     'входы в штрафную', None, None),
    ('attack', 'touchesInPenArea', 'касания мяча в штрафной', None, None),
    # ── attack: possession / losses ──
    ('attack', 'lostBall',                 'потери мяча', None, None),
    ('attack', 'loseOnOwnHalf',            'потери на своей половине', None, None),
    ('attack', 'dangerousLosesOnOwnHalf',  'опасные потери на своей половине', None, None),
    ('attack', 'technicalMistake',         'технические ошибки', None, None),
    ('attack', 'foulsSuffered',            'фолы на игроке', None, None),
    # ── attack: dribbling ──
    ('attack', 'dribble',          'дриблинг', 'удачный дриблинг', 'успешность дриблинга'),
    # ── attack: shooting ──
    ('attack', 'shot',             'удары', 'точные удары', 'точность ударов'),
    ('attack', 'byHead',           'удары головой', 'точные удары головой', 'точность ударов головой'),
    ('attack', 'goal',             'голы', None, None),
    ('attack', 'ownGoal',          'автоголы', None, None),
    ('attack', 'offside',          'офсайды', None, None),
    # ── attack: set pieces ──
    ('attack', 'penalty',          'пенальти', 'удачные пенальти', 'точность пенальти'),
    ('attack', 'freeKick',         'штрафные удары', 'удачные штрафные удары', 'точность штрафных ударов'),
    ('attack', 'directFreeKick',   'прямые штрафные', 'удачные прямые штрафные', 'точность прямых штрафных'),
    ('attack', 'corner',           'угловые', 'удачные угловые', 'точность угловых'),
    ('attack', 'throwing',         'ауты', 'удачные ауты', 'точность аутов'),
    # ── defence: tackling / recoveries ──
    ('defence', 'tackle',          'отборы', 'удачные отборы', 'успешность отборов'),
    ('defence', 'slidingTackles',  'подкаты', 'удачные подкаты', 'успешность подкатов'),
    ('defence', 'dribbleAgainst',  'дриблинги против', 'удачные дриблинги против', 'успешность дриблингов против'),
    ('defence', 'recovery',        'возвраты', None, None),
    ('defence', 'recoveryOpp',     'возвраты на чужой половине', None, None),
    ('defence', 'interception',    'перехваты', None, None),
    ('defence', 'rebounds',        'подборы', None, None),
    ('defence', 'reboundsOpp',     'подборы на чужой половине', None, None),
    # ── defence: positioning ──
    ('defence', 'clearance',       'выносы', 'точные выносы', 'точность выносов'),
    ('defence', 'blockedShot',     'заблокированные удары', None, None),
    # ── defence: duels ──
    ('defence', 'duel',            'дуэли', 'выигранные дуэли', 'успешность дуэлей'),
    ('defence', 'aerialDuel',      'воздушные дуэли', 'выигранные воздушные дуэли', 'успешность воздушных дуэлей'),
    # ── defence: pressing / discipline ──
    ('defence', 'pressing',        'прессинг', None, None),
    ('defence', 'counterpressing', 'контрпрессинг', None, None),
    ('defence', 'fouls',           'нарушения', None, None),
    ('defence', 'yellowCards',     'жёлтые карточки', None, None),
    ('defence', 'redCards',        'красные карточки', None, None),
    # ── defence: goalkeeping ──
    ('defence', 'save',            'сейвы', None, None),
    ('defence', 'goalkeeperExits', 'выходы вратаря', 'удачные выходы вратаря', 'точность выходов вратаря'),
    ('defence', 'shotsAgainst',    'удары против', 'удары против в створ', 'точность ударов против'),
    ('defence', 'goalKick',        'удары от ворот', 'удачные удары от ворот', 'точность ударов от ворот'),
    # ── fitness ──
    ('fitness', 'totalDistance',   'общая дистанция', None, None),
    ('fitness', 'speed_4_5_5',     'дистанции со скоростью 4-5.5м/с', None, None),
    ('fitness', 'speed_5_5_7',     'дистанции со скоростью 5.5-7м/с', None, None),
    ('fitness', 'speed_7plus',     'дистанции со скоростью > 7м/с', None, None),
    ('fitness', 'sprintDistance',  'дистанция спринтов', None, None),
    ('fitness', 'sprintsCount',    'количество спринтов', None, None),
    ('fitness', 'averageSpeed',    'средняя скорость', None, None),
    ('fitness', 'intenseRunning',  'интенсивный бег %', None, None),
    ('fitness', 'sprints',         'спринты', None, None),
    ('fitness', 'accelerations',   'ускорения', None, None),
    ('fitness', 'progressiveRuns', 'прогрессивные забегания', None, None),
]


# Русское имя колонки → английское (вариант экспорта SportVisor с англ.
# заголовками: «pass», «shot», «tackle»…). Опечатки экспорта сохранены
# (throught_pass, recieved_pass, ariel_duel, contrpressing). Резолвер пробует
# сначала русскую колонку, затем английскую — один парсер на оба формата.
COL_RU_EN = {
    'пас': 'pass', 'удачные пасы': 'pass_success', 'точность пасов': 'pass_accuracy',
    'пасы вперед': 'pass_forward', 'удачные пасы вперед': 'pass_forward_success', 'точность пасов вперед': 'pass_forward_accuracy',
    'пасы назад': 'pass_back', 'удачные пасы назад': 'pass_back_success', 'точность пасов назад': 'pass_back_accuracy',
    'пасы поперек': 'pass_sideways', 'удачные пасы поперек': 'pass_sideways_success', 'точность пасов поперек': 'pass_sideways_accuracy',
    'короткие пасы': 'short_pass', 'удачные короткие пасы': 'short_pass_success', 'точность коротких пасов': 'short_pass_accuracy',
    'средние пасы': 'middle_pass', 'удачные средние пасы': 'middle_pass_success', 'точность средних пасов': 'middle_pass_accuracy',
    'длинные пасы': 'long_pass', 'удачные длинные пасы': 'long_pass_success', 'точность длинных пасов': 'long_pass_accuracy',
    'пасы в штрафную': 'pass_in_pen_area', 'удачные пасы в штрафную': 'pass_in_pen_area_success', 'точность пасов в штрафную': 'pass_in_pen_area_accuracy',
    'пасы в финальную треть': 'pass_to_final_third', 'удачные пасы в финальную треть': 'pass_to_final_third_success', 'точность пасов в финальную треть': 'pass_to_final_third_accuracy',
    'прогрессивные пасы': 'progressive_pass', 'удачные прогрессивные пасы': 'progressive_pass_success', 'точность прогрессивных пасов': 'progressive_pass_accuracy',
    'разрезающие пасы': 'throught_pass', 'удачные разрезающие пасы': 'throught_pass_success', 'точность разрезающих пасов': 'throught_pass_accuracy',
    'ключевые пасы': 'key_pass',
    'кроссы': 'cross', 'удачные кроссы': 'cross_success', 'точность кроссов': 'cross_accuracy',
    'голевые передачи': 'assist', 'предголевые передачи': 'second_assist', 'третьи передачи (third assist)': 'third_assist',
    'передачи под удар в створ': 'shot_on_target_assist', 'принятые пасы': 'recieved_pass',
    'входы в штрафную': 'entries_in_penalty_area', 'касания мяча в штрафной': 'ball_touch_pen_area',
    'потери мяча': 'lost_ball', 'потери на своей половине': 'lost_ball_own_half', 'опасные потери на своей половине': 'danger_lost_ball_own_half',
    'технические ошибки': 'technical_mistake', 'фолы на игроке': 'foul_suffered',
    'дриблинг': 'dribble', 'удачный дриблинг': 'dribble_success', 'успешность дриблинга': 'dribble_accuracy',
    'удары': 'shot', 'точные удары': 'shot_success', 'точность ударов': 'shot_accuracy',
    'удары головой': 'headshot', 'точные удары головой': 'headshot_success', 'точность ударов головой': 'headshot_accuracy',
    'голы': 'goal', 'автоголы': 'auto_goal', 'офсайды': 'offside',
    'пенальти': 'penalty', 'удачные пенальти': 'penalty_success', 'точность пенальти': 'penalty_accuracy',
    'штрафные удары': 'free_kick', 'удачные штрафные удары': 'free_kick_success', 'точность штрафных ударов': 'free_kick_accuracy',
    'прямые штрафные': 'direct_freekick', 'удачные прямые штрафные': 'direct_freekick_success', 'точность прямых штрафных': 'direct_freekick_accuracy',
    'угловые': 'corner', 'удачные угловые': 'corner_success', 'точность угловых': 'corner_accuracy',
    'ауты': 'throwing', 'удачные ауты': 'throwing_success', 'точность аутов': 'throwing_accuracy',
    'отборы': 'tackle', 'удачные отборы': 'tackle_success', 'успешность отборов': 'tackle_accuracy',
    'подкаты': 'sliding_tackle', 'удачные подкаты': 'sliding_tackle_success', 'успешность подкатов': 'sliding_tackle_accuracy',
    'дриблинги против': 'dribble_against', 'удачные дриблинги против': 'dribble_against_success', 'успешность дриблингов против': 'dribble_against_accuracy',
    'возвраты': 'recovery', 'возвраты на чужой половине': 'recovery_in_opp_half', 'перехваты': 'interception',
    'выносы': 'clearance', 'точные выносы': 'clearance_success', 'точность выносов': 'clearance_accuracy', 'заблокированные удары': 'blocked_shot',
    'дуэли': 'duel', 'выигранные дуэли': 'duel_success', 'успешность дуэлей': 'duel_accuracy',
    'воздушные дуэли': 'ariel_duel', 'выигранные воздушные дуэли': 'ariel_duel_success', 'успешность воздушных дуэлей': 'ariel_duel_accuracy',
    'прессинг': 'pressing', 'контрпрессинг': 'contrpressing',
    'нарушения': 'foul', 'жёлтые карточки': 'yellow_card', 'красные карточки': 'red_card', 'сейвы': 'save',
    'выходы вратаря': 'goalkeeper_exit', 'удачные выходы вратаря': 'goalkeeper_exit_success', 'точность выходов вратаря': 'goalkeeper_exit_accuracy',
    'удары против': 'shot_against', 'удары против в створ': 'shot_against_success', 'точность ударов против': 'shot_against_accuracy',
    'удары от ворот': 'goal_kick', 'удачные удары от ворот': 'goal_kick_success', 'точность ударов от ворот': 'goal_kick_accuracy',
    'общая дистанция': 'distance_total',
    'дистанции со скоростью 4-5.5м/с': 'distance 4-5.5m/s', 'дистанции со скоростью 5.5-7м/с': 'distance 5.5-7m/s', 'дистанции со скоростью > 7м/с': 'distance > 7m/s',
    'дистанция спринтов': 'sprints_distance', 'количество спринтов': 'sprints_count', 'средняя скорость': 'speed_average', 'интенсивный бег %': 'intense_running %',
    'спринты': 'sprint', 'ускорения': 'acceleration', 'прогрессивные забегания': 'progressive_run',
}


def _resolve(row, col):
    """Значение колонки по русскому имени, иначе по английскому эквиваленту."""
    if col is None:
        return None
    v = row.get(col)
    if v is None:
        en = COL_RU_EN.get(col)
        if en:
            v = row.get(en)
    return v


def to_num(s):
    """'46' → 46, '80.4' → 80.4, '' → None. Дробное с .0 → int."""
    if s is None:
        return None
    s = str(s).strip().replace(',', '.')
    if s == '' or s == '-':
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f == int(f) else round(f, 2)


def build_value(row, total_col, succ_col, acc_col):
    """Скаляр или компаунд {total, successful, accuracy} по тройке колонок.
    Колонки резолвятся рус→англ (двуязычный экспорт)."""
    total = to_num(_resolve(row, total_col))
    if total is None:
        return None
    if succ_col is None and acc_col is None:
        return total
    succ = to_num(_resolve(row, succ_col)) if succ_col else None
    acc = to_num(_resolve(row, acc_col)) if acc_col else None
    out = {'total': total}
    if succ is not None:
        out['successful'] = succ
    if acc is not None:
        out['accuracy'] = acc
    # Если ни succ, ни acc не пришли — отдаём просто число (как PDF для скаляров).
    return out if len(out) > 1 else total


def _first(row, *cols):
    for c in cols:
        v = row.get(c)
        if v is not None and str(v).strip() != '':
            return v
    return None


def _read_input(path):
    """→ (meta, headers, rows[dict]). CSV (#-метаданные) ИЛИ xlsx (метаданные в
    первых строках «Team:/Home team:/Guest Team:/Date:», шапка с «номер»/«jersey no.»)."""
    meta = {}
    if str(path).lower().endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb['Player Report'] if 'Player Report' in wb.sheetnames else wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        hi = None
        for i, r in enumerate(grid):
            cells = [str(c).strip().lower() if c is not None else '' for c in r]
            if 'номер' in cells or 'jersey no.' in cells:
                hi = i
                break
            c0 = str(r[0]).strip() if r and r[0] is not None else ''
            m = re.match(r'^([^:]+):\s*(.+)$', c0)
            if m:
                meta[m.group(1).strip().lower()] = m.group(2).strip()
        if hi is None:
            return meta, [], []
        headers = [str(c).strip() if c is not None else '' for c in grid[hi]]
        rows = [dict(zip(headers, list(r) + [None] * (len(headers) - len(r)))) for r in grid[hi + 1:]]
        return meta, headers, rows
    with open(path, encoding='utf-8-sig') as f:
        lines = f.readlines()
    data_lines = []
    for ln in lines:
        if ln.startswith('#'):
            m = re.match(r'^#\s*([^:]+):\s*(.+)$', ln.strip())
            if m:
                meta[m.group(1).strip().lower()] = m.group(2).strip()
        else:
            data_lines.append(ln)
    reader = csv.DictReader(io.StringIO(''.join(data_lines)))
    return meta, (reader.fieldnames or []), list(reader)


def parse(csv_path):
    meta, headers, rows = _read_input(csv_path)

    players = []
    for row in rows:
        num_raw = _first(row, 'номер', 'jersey no.')
        if num_raw is None:
            continue
        try:
            number = str(int(float(num_raw))).zfill(2)
        except (ValueError, TypeError):
            continue
        name = str(_first(row, 'имя', 'name') or '').strip()
        minutes = to_num(_first(row, 'игровое время', 'playing_time'))
        if isinstance(minutes, float):
            minutes = int(minutes)

        stats = {'attack': {}, 'defence': {}, 'fitness': {}}
        for group, key, tcol, scol, acol in MAPPING:
            v = build_value(row, tcol, scol, acol)
            if v is not None:
                stats[group][key] = v

        skip = ('номер', 'имя', 'jersey no.', 'name')
        players.append({
            'number': number,
            'name': name,
            'minutes': minutes,
            'stats': stats,
            'raw': {h: row.get(h) for h in headers if h not in skip},
        })

    score = None
    for k in ('result', 'score', 'счет', 'счёт'):
        if k in meta:
            mm = re.match(r'(\d+)\s*[:-]\s*(\d+)', meta[k])
            if mm:
                score = {'home': int(mm.group(1)), 'away': int(mm.group(2))}
                break

    return {
        'match': {
            'homeTeam':     meta.get('home team') or meta.get('team'),
            'awayTeam':     meta.get('guest team'),
            'date':         meta.get('date'),
            'score':        score,
            'source':       'csv',
            'columnsCount': len(headers),
        },
        'players': players,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('input_csv')
    ap.add_argument('output_json')
    args = ap.parse_args()
    data = parse(args.input_csv)
    with open(args.output_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    if not data['players']:
        print('WARN: 0 players parsed — проверь, что в CSV есть колонка «номер»', file=sys.stderr)
    print(f"OK columns={data['match']['columnsCount']} players={len(data['players'])}", file=sys.stderr)


if __name__ == '__main__':
    main()
